## Code for reading an excel (xlxs) file 1st column and move/copy those images from 1 folder to other

from openpyxl import load_workbook
from tqdm import tqdm

#give your excel sheet address with name
wb = load_workbook("D:/sheets/tmp.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))]
#print(column_list, ' column items')

#### now for copy/moving the images

import shutil, os

for f in tqdm(column_list):

    #1st argument is origin from where you want to copy and second is destination
    #shutil.move('D:/images1/' + str(f)+'.png', 'D:/images2/')
    shutil.copy('D:/images1/' + str(f)+'.png', 'D:/images2/')

# 	counter += 1
# 	print(counter, ' ', f, ' copied')





